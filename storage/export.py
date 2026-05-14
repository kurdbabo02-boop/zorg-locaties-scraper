"""
Export zorglocaties naar CSV en Excel.
Overzichtelijk: alleen naam, locatie, telefoon, e-mail en website.
"""

import csv
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import List

from models import CareLocation
from config.settings import OUTPUT_DIR

logger = logging.getLogger(__name__)

# De 5 kolommen die we tonen
CLEAN_FIELDS = ["naam", "locatie", "telefoon", "email", "website"]


def _ensure_dir():
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M")


def _to_clean_row(loc: CareLocation) -> dict:
    """Zet een CareLocation om naar de 5 overzichtelijke kolommen."""
    # Locatie samenstellen: "Straatnaam 1, 1234 AB Amsterdam (NL)"
    parts = []
    if loc.address:
        parts.append(loc.address)
    if loc.postal_code and loc.city:
        parts.append(f"{loc.postal_code} {loc.city}")
    elif loc.city:
        parts.append(loc.city)
    if loc.country:
        parts.append(f"({loc.country})")
    locatie = ", ".join(parts) if parts else ""

    return {
        "naam":     loc.name,
        "locatie":  locatie,
        "telefoon": loc.phone,
        "email":    loc.email,
        "website":  loc.website or loc.source_url,
    }


def to_csv(locations: List[CareLocation], filename: str = None) -> str:
    _ensure_dir()
    if not filename:
        filename = os.path.join(OUTPUT_DIR, f"zorg_locaties_{_timestamp()}.csv")

    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CLEAN_FIELDS)
        writer.writeheader()
        for loc in locations:
            writer.writerow(_to_clean_row(loc))

    logger.info("CSV opgeslagen: %s (%d rijen)", filename, len(locations))
    return filename


def to_excel(locations: List[CareLocation], filename: str = None) -> str:
    """Netjes opgemaakte Excel met de 5 overzichtskolommen."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl niet geïnstalleerd — Excel export overgeslagen.")
        return ""

    _ensure_dir()
    if not filename:
        filename = os.path.join(OUTPUT_DIR, f"zorg_locaties_{_timestamp()}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zorg Locaties"

    # ---- Stijlen ----
    HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")    # donkerblauw
    ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")    # lichtblauw
    CELL_FONT    = Font(name="Calibri", size=10)
    LINK_FONT    = Font(name="Calibri", size=10, color="0563C1", underline="single")
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    THIN_BORDER  = Border(
        bottom=Side(style="thin", color="D9D9D9"),
    )

    COLUMN_WIDTHS = {
        "naam":     35,
        "locatie":  40,
        "telefoon": 18,
        "email":    30,
        "website":  45,
    }
    COLUMN_LABELS = {
        "naam":     "Naam instelling",
        "locatie":  "Locatie",
        "telefoon": "Telefoon",
        "email":    "E-mail",
        "website":  "Website",
    }

    # ---- Header rij ----
    ws.row_dimensions[1].height = 22
    for col_idx, field in enumerate(CLEAN_FIELDS, 1):
        cell = ws.cell(row=1, column=col_idx, value=COLUMN_LABELS[field])
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    # ---- Data rijen ----
    for row_idx, loc in enumerate(locations, 2):
        row = _to_clean_row(loc)
        fill = ALT_FILL if row_idx % 2 == 0 else None
        ws.row_dimensions[row_idx].height = 16

        for col_idx, field in enumerate(CLEAN_FIELDS, 1):
            value = row[field] or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = LEFT
            cell.border    = THIN_BORDER
            if fill:
                cell.fill = fill

            # Website en e-mail als klikbare hyperlink
            if field == "website" and value.startswith("http"):
                cell.hyperlink = value
                cell.font = LINK_FONT
            elif field == "email" and "@" in value:
                cell.hyperlink = f"mailto:{value}"
                cell.font = LINK_FONT
            else:
                cell.font = CELL_FONT

    # ---- Kolombreedtes ----
    for col_idx, field in enumerate(CLEAN_FIELDS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[field]

    # ---- Filters + bevroren header ----
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CLEAN_FIELDS))}1"
    ws.freeze_panes = "A2"

    wb.save(filename)
    logger.info("Excel opgeslagen: %s (%d rijen)", filename, len(locations))
    return filename


def to_json(locations: List[CareLocation], filename: str = None) -> str:
    import json
    _ensure_dir()
    if not filename:
        filename = os.path.join(OUTPUT_DIR, f"zorg_locaties_{_timestamp()}.json")
    data = [_to_clean_row(loc) for loc in locations]
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.info("JSON opgeslagen: %s (%d items)", filename, len(data))
    return filename


def export_all(locations: List[CareLocation]) -> dict:
    ts = _timestamp()
    base = os.path.join(OUTPUT_DIR, f"zorg_locaties_{ts}")
    return {
        "csv":   to_csv(locations,   base + ".csv"),
        "json":  to_json(locations,  base + ".json"),
        "excel": to_excel(locations, base + ".xlsx"),
    }
