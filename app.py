"""
Zorg Locaties — Streamlit interface
Vindt 24/7 zorginstellingen in Nederland en België.
"""

import io, sys, os, smtplib, ssl
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from typing import List

import pandas as pd
import streamlit as st

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from models import CareLocation
from config.regions import NL_PROVINCES, BE_PROVINCES

# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Zorg Locaties Finder",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Sessiestatus ──────────────────────────────────────────────────────────────
DEFAULTS = dict(results=[], last_run=None, df=None, dark_mode=False,
                email_sent=[], smtp_ok=False)
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Thema CSS ─────────────────────────────────────────────────────────────────
LIGHT = """
<style>
:root { --bg:#ffffff; --bg2:#f0f6ff; --sidebar:#f8faff;
        --text:#1a1a1a; --accent:#1F4E79; --accent2:#163d61;
        --card-bg:#f0f6ff; --border:#d0dff0; --muted:#888; }
</style>"""

DARK = """
<style>
:root { --bg:#0e1117; --bg2:#1a1f2e; --sidebar:#151b28;
        --text:#f0f4ff; --accent:#4da6ff; --accent2:#6bb8ff;
        --card-bg:#1e2535; --border:#2d3a52; --muted:#8a9bbf; }
.stApp { background-color: #0e1117 !important; color: #f0f4ff !important; }
section[data-testid="stSidebar"] { background-color: #151b28 !important; }
section[data-testid="stSidebar"] * { color: #f0f4ff !important; }
.stTextInput input, .stTextArea textarea, .stSelectbox div,
.stMultiSelect div { background-color: #1e2535 !important; color: #f0f4ff !important; border-color: #2d3a52 !important; }
.stDataFrame, [data-testid="stDataFrame"] { background-color: #1e2535 !important; }
h1,h2,h3,h4,h5,h6,p,label,span,.stMarkdown { color: #f0f4ff !important; }
.stButton button { background-color: #1e2535 !important; color: #f0f4ff !important; border-color: #2d3a52 !important; }
</style>"""

COMMON_CSS = """
<style>
    .main-title { font-size:2rem; font-weight:700; color:var(--accent); margin-bottom:0; }
    .sub-title  { font-size:.9rem; color:var(--muted); margin-top:0; margin-bottom:1.2rem; }
    .stat-card  { background:var(--card-bg); border-left:4px solid var(--accent);
                  border-radius:6px; padding:12px 16px; margin-bottom:8px; }
    .stat-label { font-size:.7rem; color:var(--muted); text-transform:uppercase; letter-spacing:.05em; }
    .stat-value { font-size:1.6rem; font-weight:700; color:var(--accent); line-height:1.2; }
    .sidebar-section { font-size:.7rem; font-weight:700; color:var(--accent);
                       text-transform:uppercase; letter-spacing:.07em;
                       margin-top:.9rem; margin-bottom:.2rem; }
    .sel-badge { background:var(--accent); color:white; border-radius:12px;
                 padding:3px 10px; font-size:.8rem; font-weight:600;
                 display:inline-block; margin-bottom:.5rem; }
    .email-preview { background:var(--card-bg); border:1px solid var(--border);
                     border-radius:8px; padding:16px; font-family:monospace;
                     font-size:.85rem; white-space:pre-wrap; }
</style>"""

st.markdown(DARK if st.session_state.dark_mode else LIGHT, unsafe_allow_html=True)
st.markdown(COMMON_CSS, unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🏥 Zorg Locaties")
    st.caption("24/7 zorginstellingen in NL & BE")

    # Thema toggle
    dark_on = st.toggle("🌙 Dark mode", value=st.session_state.dark_mode)
    if dark_on != st.session_state.dark_mode:
        st.session_state.dark_mode = dark_on
        st.rerun()

    st.divider()

    st.markdown('<div class="sidebar-section">🌍 Land</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    include_nl = c1.checkbox("🇳🇱 NL", value=True)
    include_be = c2.checkbox("🇧🇪 BE", value=True)

    st.markdown('<div class="sidebar-section">📍 Regio</div>', unsafe_allow_html=True)
    region_pool = (NL_PROVINCES if include_nl else []) + (BE_PROVINCES if include_be else [])
    selected_regions = st.multiselect(
        "Provincies (leeg = heel NL + BE)",
        options=sorted(set(region_pool)), placeholder="Alle provincies",
    )

    st.markdown('<div class="sidebar-section">🏥 Zorgtype (24/7)</div>', unsafe_allow_html=True)
    care_types = st.multiselect(
        "Type instelling",
        options=["Verpleeghuis","Woonzorgcentrum","Kleinschalig wonen",
                 "Alzheimer / dementie centrum","Hospice"],
        default=["Verpleeghuis","Woonzorgcentrum","Kleinschalig wonen",
                 "Alzheimer / dementie centrum"],
    )

    st.markdown('<div class="sidebar-section">🎯 Specialisatie</div>', unsafe_allow_html=True)
    focus_dem = st.checkbox("🧠 Dementie / Alzheimer", value=True)
    focus_eld = st.checkbox("👴 Ouderenzorg algemeen", value=True)

    st.markdown('<div class="sidebar-section">🔍 Focus</div>', unsafe_allow_html=True)
    only_small    = st.checkbox("Alleen kleine instellingen")
    only_emerging = st.checkbox("Alleen opkomende / nieuwe")

    st.markdown('<div class="sidebar-section">⚡ Zoeksnelheid</div>', unsafe_allow_html=True)
    quick_mode = st.toggle("Snel zoeken (aanbevolen)", value=True,
        help="Snel: ~8 queries, <30 sec.\nUitgebreid: alle regio's, kan minuten duren.")

    st.markdown('<div class="sidebar-section">⚙️ Bronnen</div>', unsafe_allow_html=True)
    use_zorgkaart = st.checkbox("Zorgkaart NL",          value=True)
    use_vektis    = st.checkbox("Vektis AGB open data",  value=True)
    use_search    = st.checkbox("DuckDuckGo zoeken",     value=True)
    use_belgium   = st.checkbox("Belgische bronnen",     value=include_be)

    st.divider()
    start_btn = st.button("🔍 Start zoeken", type="primary",
                          use_container_width=True,
                          disabled=not (include_nl or include_be))

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown('<p class="main-title">🏥 Zorg Locaties Finder</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">24/7 zorginstellingen in Nederland en België — ouderenzorg & dementie</p>',
            unsafe_allow_html=True)

# ── Constanten ────────────────────────────────────────────────────────────────
CARE_TYPE_MAP = {
    "Verpleeghuis":                 ["verpleeghuis","ouderengeneeskunde"],
    "Woonzorgcentrum":              ["woonzorgcentrum"],
    "Kleinschalig wonen":           ["kleinschalig_wonen"],
    "Alzheimer / dementie centrum": ["dementie_centrum","alzheimer_centrum"],
    "Hospice":                      ["hospice"],
}
EXCL_TERMS = ["thuiszorg","dagopvang","dagverzorging","thuisverpleging",
              "vacature","werken bij","solliciteer"]

# ── Scrapers uitvoeren ────────────────────────────────────────────────────────
def run_scrapers_ui(countries, regions, care_sel, f_dem, f_eld,
                    sm, em, flags, quick) -> List[CareLocation]:
    from scrapers import ZorgkaartScraper, SearchScraper, VektisScraper, BelgiumScraper
    from config.queries import EXCLUDED_CARE_TYPES

    locs: List[CareLocation] = []
    if flags["zorgkaart"] and "NL" in countries:
        locs.extend(ZorgkaartScraper().scrape())
    if flags["vektis"] and "NL" in countries:
        locs.extend(VektisScraper().scrape())
    if flags["search"]:
        locs.extend(SearchScraper(quick_mode=quick, region_override=regions or None).scrape())
    if flags["belgium"] and "BE" in countries:
        locs.extend(BelgiumScraper().scrape())

    locs = [l for l in locs if l.care_type not in EXCLUDED_CARE_TYPES]
    locs = [l for l in locs if not any(t in l.name.lower() for t in EXCL_TERMS)]
    locs = [l for l in locs if l.country in countries]
    if regions:
        locs = [l for l in locs if not l.province or
                any(r.lower() in (l.province+" "+l.city).lower() for r in regions)]
    allowed = []
    for label in care_sel:
        allowed.extend(CARE_TYPE_MAP.get(label, []))
    if allowed:
        locs = [l for l in locs if not l.care_type or l.care_type in allowed]
    if f_dem and not f_eld:
        locs = [l for l in locs if "dementie" in l.specializations]
    elif f_eld and not f_dem:
        locs = [l for l in locs if "ouderenzorg" in l.specializations]
    if sm: locs = [l for l in locs if l.is_small]
    if em: locs = [l for l in locs if l.is_emerging]

    seen, unique = set(), []
    for l in locs:
        k = l.dedup_key()
        if k not in seen:
            seen.add(k); unique.append(l)
    return unique


def make_df(locs: List[CareLocation]) -> pd.DataFrame:
    rows = []
    for l in locs:
        parts = [p for p in [l.address, f"{l.postal_code} {l.city}".strip(),
                              f"({l.country})"] if p.strip()]
        rows.append({
            "☑":        False,
            "Naam":     l.name,
            "Locatie":  ", ".join(parts),
            "Type":     l.care_type.replace("_"," ").title() if l.care_type else "",
            "Telefoon": l.phone or "",
            "E-mail":   l.email or "",
            "Website":  l.website or l.source_url or "",
            "Klein":    "✓" if l.is_small else "",
            "Nieuw":    "✓" if l.is_emerging else "",
        })
    return pd.DataFrame(rows)


# ── Starten ───────────────────────────────────────────────────────────────────
if start_btn:
    countries = [c for c, inc in [("NL",include_nl),("BE",include_be)] if inc]
    flags = dict(zorgkaart=use_zorgkaart, vektis=use_vektis,
                 search=use_search, belgium=use_belgium)
    steps = [k for k, v in flags.items() if v]
    bar   = st.progress(0, text="Initialiseren...")
    info  = st.empty()
    all_locs: List[CareLocation] = []

    LABELS = dict(zorgkaart="Zorgkaart Nederland", vektis="Vektis AGB-register",
                  search="DuckDuckGo zoeken", belgium="Belgische bronnen")

    for i, step in enumerate(steps):
        info.info(f"🔍 {LABELS[step]}...")
        bar.progress(i/len(steps), text=f"Bezig: {LABELS[step]}...")
        try:
            partial = run_scrapers_ui(
                countries, selected_regions, care_types,
                focus_dem, focus_eld, only_small, only_emerging,
                {k:(k==step) for k in flags}, quick_mode,
            )
            all_locs.extend(partial)
        except Exception as e:
            st.warning(f"⚠️ {LABELS[step]} mislukt: {e}")

    bar.progress(1.0, text="Klaar!")
    info.empty()

    seen, unique = set(), []
    for l in all_locs:
        k = l.dedup_key()
        if k not in seen:
            seen.add(k); unique.append(l)

    st.session_state.results  = unique
    st.session_state.last_run = datetime.now().strftime("%d-%m-%Y %H:%M")
    st.session_state.df       = make_df(unique)
    st.session_state.email_sent = []

# ── Resultaten ────────────────────────────────────────────────────────────────
results: List[CareLocation] = st.session_state.results

if results:
    df_master: pd.DataFrame = st.session_state.df

    nl = sum(1 for l in results if l.country=="NL")
    be = sum(1 for l in results if l.country=="BE")
    sm = sum(1 for l in results if l.is_small)
    c1,c2,c3,c4 = st.columns(4)
    for col, lbl, val in [(c1,"Totaal",len(results)),(c2,"Nederland",nl),
                          (c3,"België",be),(c4,"Kleinschalig",sm)]:
        col.markdown(
            f'<div class="stat-card"><div class="stat-label">{lbl}</div>'
            f'<div class="stat-value">{val}</div></div>', unsafe_allow_html=True)
    if st.session_state.last_run:
        st.caption(f"Laatste zoekopdracht: {st.session_state.last_run}")

    st.divider()

    search_q = st.text_input("🔎 Zoek in resultaten",
                             placeholder="bijv. Amsterdam of verpleeghuis...")
    df_view = df_master.copy()
    if search_q:
        mask = (df_view["Naam"].str.contains(search_q,case=False,na=False) |
                df_view["Locatie"].str.contains(search_q,case=False,na=False) |
                df_view["Type"].str.contains(search_q,case=False,na=False))
        df_view = df_view[mask].reset_index(drop=True)

    st.caption(f"{len(df_view)} van {len(results)} locaties — **klik ☑ om rijen te selecteren**")

    edited = st.data_editor(
        df_view,
        use_container_width=True,
        hide_index=True,
        height=460,
        column_config={
            "☑":        st.column_config.CheckboxColumn("☑",       width="small", default=False),
            "Naam":     st.column_config.TextColumn("Naam",         width="large"),
            "Locatie":  st.column_config.TextColumn("Locatie",      width="large"),
            "Type":     st.column_config.TextColumn("Type",         width="medium"),
            "Telefoon": st.column_config.TextColumn("Telefoon",     width="medium"),
            "E-mail":   st.column_config.TextColumn("E-mail",       width="medium"),
            "Website":  st.column_config.LinkColumn("Website", display_text="🔗 Bekijk", width="small"),
            "Klein":    st.column_config.TextColumn("Klein",        width="small"),
            "Nieuw":    st.column_config.TextColumn("Nieuw",        width="small"),
        },
        disabled=["Naam","Locatie","Type","Telefoon","Website","Klein","Nieuw"],
        key="table_editor",
    )

    selected_df = edited[edited["☑"] == True].copy()
    n_sel = len(selected_df)
    if n_sel > 0:
        st.markdown(f'<div class="sel-badge">✓ {n_sel} rijen geselecteerd</div>',
                    unsafe_allow_html=True)

    # ── Downloads ─────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 📥 Downloaden")

    scope = st.radio("Exporteer:", ["Alle resultaten","Alleen geselecteerde rijen"],
                     horizontal=True, disabled=(n_sel==0),
                     index=1 if n_sel>0 else 0)
    export_src = selected_df if (scope=="Alleen geselecteerde rijen" and n_sel>0) else edited
    COLS = ["Naam","Locatie","Type","Telefoon","E-mail","Website"]
    ex_df = export_src[COLS].copy()
    st.caption(f"{len(ex_df)} rijen worden geëxporteerd")

    dl1, dl2 = st.columns(2)
    with dl1:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            xbuf = io.BytesIO()
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Zorg Locaties"
            HF = Font(name="Calibri",bold=True,color="FFFFFF",size=11)
            HFI = PatternFill("solid",fgColor="1F4E79")
            AF  = PatternFill("solid",fgColor="EBF3FB")
            CF  = Font(name="Calibri",size=10)
            LF  = Font(name="Calibri",size=10,color="0563C1",underline="single")
            AL  = Alignment(horizontal="left",vertical="center")
            BO  = Border(bottom=Side(style="thin",color="D9D9D9"))
            WS  = [38,42,22,18,32,45]
            LB  = ["Naam instelling","Locatie","Type","Telefoon","E-mail","Website"]
            ws.row_dimensions[1].height = 22
            for ci,(lb,w) in enumerate(zip(LB,WS),1):
                c = ws.cell(row=1,column=ci,value=lb)
                c.font=HF; c.fill=HFI
                c.alignment=Alignment(horizontal="center",vertical="center")
                ws.column_dimensions[get_column_letter(ci)].width=w
            for ri,(_,row) in enumerate(ex_df.iterrows(),2):
                fill=AF if ri%2==0 else None
                ws.row_dimensions[ri].height=16
                for ci,col in enumerate(COLS,1):
                    val=str(row.get(col,"") or "")
                    cell=ws.cell(row=ri,column=ci,value=val)
                    cell.alignment=AL; cell.border=BO
                    if fill: cell.fill=fill
                    if col=="Website" and val.startswith("http"):
                        cell.hyperlink=val; cell.value="Bekijk website"; cell.font=LF
                    elif col=="E-mail" and "@" in val:
                        cell.hyperlink=f"mailto:{val}"; cell.font=LF
                    else:
                        cell.font=CF
            ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}1"
            ws.freeze_panes="A2"
            wb.save(xbuf); xbuf.seek(0)
            dl1.download_button("📊 Download Excel", data=xbuf,
                file_name=f"zorg_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        except Exception as e:
            dl1.error(f"Excel fout: {e}")

    with dl2:
        csv_b = ex_df.to_csv(index=False,encoding="utf-8-sig").encode("utf-8-sig")
        dl2.download_button("📄 Download CSV", data=csv_b,
            file_name=f"zorg_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv", use_container_width=True)

    # ── E-mail versturen ──────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 📧 E-mails versturen")

    if n_sel == 0:
        st.info("Selecteer eerst rijen in de tabel (☑) om e-mails te versturen.")
    else:
        # Tabel van ontvangers met bewerkbare e-mail kolom
        st.markdown(f"**{n_sel} instellingen geselecteerd** — vul ontbrekende e-mailadressen in:")
        recip_df = selected_df[["Naam","E-mail"]].copy().reset_index(drop=True)
        recip_edited = st.data_editor(
            recip_df,
            use_container_width=True,
            hide_index=True,
            height=min(35 + n_sel*36, 300),
            column_config={
                "Naam":   st.column_config.TextColumn("Naam instelling", disabled=True),
                "E-mail": st.column_config.TextColumn("E-mailadres", help="Voer in als leeg"),
            },
            key="recip_editor",
        )

        valid_recips = recip_edited[recip_edited["E-mail"].str.contains("@", na=False)]
        n_valid = len(valid_recips)
        if n_valid < n_sel:
            st.caption(f"⚠️ {n_sel - n_valid} instellingen zonder e-mailadres worden overgeslagen.")

        st.markdown("**E-mailsjabloon:**")
        col_subj, col_sign = st.columns([3,1])
        with col_subj:
            subject = st.text_input("Onderwerp",
                value="Kennismaking — samenwerking ouderenzorg",
                key="email_subject")
        with col_sign:
            sender_name = st.text_input("Jouw naam", placeholder="bijv. Care", key="sender_name")

        body_template = st.text_area(
            "Berichttekst (gebruik {naam} voor de naam van de instelling)",
            value=(
                "Geachte {naam},\n\n"
                "Mijn naam is {afzender} en ik neem contact met u op in verband met "
                "een mogelijke samenwerking op het gebied van ouderenzorg en dementiezorg.\n\n"
                "Ik zou graag meer willen weten over uw instelling en de mogelijkheden "
                "voor een vrijblijvend kennismakingsgesprek.\n\n"
                "Zou u open staan voor een kort gesprek?\n\n"
                "Met vriendelijke groet,\n"
                "{afzender}"
            ),
            height=220,
            key="email_body",
        )

        # Preview
        with st.expander("👁 Preview eerste e-mail"):
            preview_naam = valid_recips.iloc[0]["Naam"] if n_valid > 0 else "Verpleeghuis Voorbeeld"
            preview = body_template.replace("{naam}", preview_naam).replace(
                "{afzender}", sender_name or "Uw naam")
            st.markdown(f"**Aan:** {valid_recips.iloc[0]['E-mail'] if n_valid>0 else 'voorbeeld@zorg.nl'}")
            st.markdown(f"**Onderwerp:** {subject}")
            st.markdown(f'<div class="email-preview">{preview}</div>', unsafe_allow_html=True)

        st.markdown("**SMTP-instellingen (Gmail aanbevolen):**")
        sm1, sm2, sm3 = st.columns([2,2,1])
        smtp_email    = sm1.text_input("Jouw e-mailadres",    key="smtp_email",
                                        placeholder="jij@gmail.com")
        smtp_password = sm2.text_input("App-wachtwoord",      key="smtp_password",
                                        type="password",
                                        placeholder="xxxx xxxx xxxx xxxx")
        smtp_host     = sm3.selectbox("Provider", ["Gmail","Outlook","Anders"], key="smtp_host")

        HOST_MAP = {"Gmail":("smtp.gmail.com",587),
                    "Outlook":("smtp.office365.com",587),
                    "Anders":("smtp.gmail.com",587)}
        host, port = HOST_MAP[smtp_host]

        if smtp_host == "Gmail":
            st.caption("Gmail: ga naar **Mijn Google-account → Beveiliging → App-wachtwoorden** "
                       "en maak een App-wachtwoord aan voor 'Mail'.")

        send_col, _ = st.columns([1,3])
        with send_col:
            send_btn = st.button(
                f"📨 Verstuur {n_valid} e-mails",
                type="primary",
                disabled=(n_valid == 0 or not smtp_email or not smtp_password),
                use_container_width=True,
            )

        if send_btn:
            if not sender_name:
                st.error("Vul eerst jouw naam in.")
            else:
                results_log = []
                progress_email = st.progress(0, text="E-mails versturen...")
                ctx = ssl.create_default_context()
                try:
                    with smtplib.SMTP(host, port) as server:
                        server.starttls(context=ctx)
                        server.login(smtp_email, smtp_password)

                        for i, (_, row) in enumerate(valid_recips.iterrows()):
                            naam  = row["Naam"]
                            email = row["E-mail"]
                            tekst = body_template.replace("{naam}", naam).replace(
                                "{afzender}", sender_name)

                            msg = MIMEMultipart("alternative")
                            msg["Subject"] = subject
                            msg["From"]    = f"{sender_name} <{smtp_email}>"
                            msg["To"]      = email
                            msg.attach(MIMEText(tekst, "plain", "utf-8"))

                            try:
                                server.sendmail(smtp_email, email, msg.as_string())
                                results_log.append(("✅", naam, email))
                            except Exception as e:
                                results_log.append(("❌", naam, f"Fout: {e}"))

                            progress_email.progress((i+1)/n_valid,
                                                    text=f"Verstuurd: {naam}...")

                except smtplib.SMTPAuthenticationError:
                    st.error("❌ Inloggen mislukt. Controleer je e-mailadres en App-wachtwoord.")
                    results_log = []
                except Exception as e:
                    st.error(f"❌ SMTP fout: {e}")
                    results_log = []

                if results_log:
                    progress_email.empty()
                    ok  = [(n,e) for s,n,e in results_log if s=="✅"]
                    err = [(n,e) for s,n,e in results_log if s=="❌"]
                    if ok:
                        st.success(f"✅ {len(ok)} e-mail(s) verstuurd!")
                    if err:
                        for n,e in err:
                            st.warning(f"⚠️ {n}: {e}")
                    st.session_state.email_sent += [e for _,_,e in results_log]

        if st.session_state.email_sent:
            st.caption(f"📬 Al verstuurd in deze sessie: {len(st.session_state.email_sent)} e-mail(s)")

else:
    st.info("👈 Stel je filters in en klik **Start zoeken**.")
    st.markdown("""
    **Wat zoekt dit systeem?**
    - 🏠 Verpleeghuizen en woonzorgcentra (24/7 zorg)
    - 🧠 Dementie- en Alzheimer-afdelingen
    - 🏡 Kleinschalige woonvormen voor ouderen
    - Zowel **Nederland** als **België**

    **Tip:** gebruik **Snel zoeken** voor resultaten in <30 seconden.
    """)
